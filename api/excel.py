"""Excel-выгрузка дашборда холдинга из снапшота кэша (R3.3 — топ-приоритет).

Собирает многолистовую книгу прямо из готового снапшота (без обращения к
Omnicomm): «Холдинг · ДЗО», «Экономика», «Скоростной режим», «Контроль ТО»,
«Качество данных». Отдаётся кнопкой в портале одним файлом на текущий срез.
"""

from __future__ import annotations

import io
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_HEAD_FILL = PatternFill("solid", fgColor="2E7D6B")
_HEAD_FONT = Font(bold=True, color="FFFFFF")
_ZEBRA = PatternFill("solid", fgColor="F1F6F4")


def _sheet(wb: Workbook, title: str, headers: list[str], rows: list[list[Any]]) -> None:
    ws = wb.create_sheet(title[:31])
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _HEAD_FILL
        cell.font = _HEAD_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for i, row in enumerate(rows):
        ws.append(row)
        if i % 2 == 1:
            for c in range(1, len(headers) + 1):
                ws.cell(row=i + 2, column=c).fill = _ZEBRA
    # ширины по контенту (с потолком)
    for c in range(1, len(headers) + 1):
        letter = get_column_letter(c)
        longest = max([len(str(headers[c - 1]))] +
                      [len(str(r[c - 1])) for r in rows if c - 1 < len(r)] or [0])
        ws.column_dimensions[letter].width = min(max(longest + 2, 10), 48)
    ws.freeze_panes = "A2"


def _flatten_orgs(nodes: list[dict], depth: int = 0, out: Optional[list] = None) -> list[dict]:
    out = [] if out is None else out
    for n in nodes or []:
        out.append({"node": n, "depth": depth})
        _flatten_orgs(n.get("children", []), depth + 1, out)
    return out


def _r1(x, nd=1):
    try:
        return round(float(x), nd)
    except (TypeError, ValueError):
        return x


def build_workbook(snapshot: dict) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)  # убрать дефолтный пустой лист

    period = (snapshot.get("period") or {}).get("label", "")

    # 1. Холдинг · ДЗО (иерархия KPI)
    org_rows = []
    for item in _flatten_orgs(snapshot.get("orgs") or []):
        n = item["node"]; k = n.get("kpi") or {}
        org_rows.append([
            "  " * item["depth"] + (n.get("name") or ""),
            n.get("type") or "", n.get("vehicle_count") or 0,
            k.get("vehicles_with_data") or 0,
            _r1(k.get("total_mileage_km")), _r1(k.get("total_fuel_l")),
            _r1(k.get("total_fuel_cost"), 0), _r1(k.get("total_idle_hours")),
            _r1((k.get("idle_hours_share") or 0) * 100),
            _r1(k.get("potential_savings"), 0),
            _r1(k.get("max_speed_kmh")),
            _r1((k.get("speeding_mileage_share") or 0) * 100),
        ])
    _sheet(wb, "Холдинг · ДЗО",
           ["Организация", "Тип", "ТС", "С данными", "Пробег, км", "Топливо, л",
            "Топливо, ₸", "Простой, ч", "Простой, %", "Потенциал экономии, ₸",
            "Макс. скорость, км/ч", "Превышения, % пробега"],
           org_rows)

    # 2. Экономика (корзины денег)
    eco = snapshot.get("economics")
    if eco:
        eco_rows = [[b.get("label"), _r1(b.get("existing_kzt"), 0),
                     _r1(b.get("potential_kzt"), 0),
                     "≈ оценка" if b.get("is_estimate") else "факт",
                     b.get("note") or ""] for b in (eco.get("buckets") or [])]
        eco_rows.append(["ИТОГО потери / потенциал",
                         _r1(eco.get("total_existing_kzt"), 0),
                         _r1(eco.get("total_potential_kzt"), 0), "", ""])
        eco_rows.append(["Адресуемые потери, ₸/мес (COI)",
                         _r1(eco.get("coi_monthly_kzt"), 0), "", "", ""])
        _sheet(wb, "Экономика",
               ["Корзина", "Потери за период, ₸", "Потенциал, ₸", "Тип", "Основание"],
               eco_rows)

    # 3. Скоростной режим (рекомендации на букве закона)
    recs = snapshot.get("recommendations") or []
    if recs:
        rec_rows = []
        for r in recs:
            rec_rows.append([
                r.get("name") or r.get("terminal_id") or "",
                r.get("severity") or r.get("st_kap_severity") or "",
                r.get("koap_article") or r.get("article") or "—",
                (r.get("text") or "").strip(),
            ])
        _sheet(wb, "Скоростной режим",
               ["ТС", "Серьёзность", "Статья КоАП", "Рекомендация"], rec_rows)

    # 4. Контроль ТО
    maint = snapshot.get("maintenance")
    if maint:
        m_rows = [[m.get("name") or m.get("terminal_id"), m.get("status"),
                   _r1(m.get("mh_since")), _r1(m.get("km_since")),
                   _r1(m.get("mh_left")), _r1(m.get("km_left")), m.get("reason")]
                  for m in (maint.get("items") or [])]
        _sheet(wb, "Контроль ТО",
               ["ТС", "Статус", "Моточасы с ТО", "Пробег с ТО, км",
                "Осталось мч", "Осталось км", "Комментарий"], m_rows)

    # 4б. Учёт шин по пробегу
    tyres = snapshot.get("tyres")
    if tyres:
        t_rows = [[t.get("name") or t.get("terminal_id"), t.get("status"),
                   _r1(t.get("km_since")), _r1(t.get("resource_km")),
                   _r1((t.get("worn_share") or 0) * 100, 0), _r1(t.get("km_left")),
                   _r1(t.get("wear_kzt"), 0), t.get("brand") or "", t.get("reason")]
                  for t in (tyres.get("items") or [])]
        _sheet(wb, "Шины",
               ["ТС", "Статус", "Пробег комплекта, км", "Ресурс, км",
                "Отхожено, %", "Осталось, км", "Износ, ₸", "Бренд", "Комментарий"], t_rows)

    # 5. Качество данных (Sensor Health)
    sh = snapshot.get("sensor_health")
    if sh:
        _STAT = {"online": "🟢 онлайн", "stale": "🟡 устарели",
                 "offline": "🔴 офлайн", "unknown": "нет записи"}
        s_rows = [[t.get("name") or t.get("terminal_id"),
                   _STAT.get(t.get("status"), t.get("status")),
                   _r1((t.get("age_seconds") or 0) / 3600, 1)]
                  for t in (sh.get("terminals") or [])]
        _sheet(wb, "Качество данных",
               ["ТС", "Статус терминала", "Давность данных, ч"], s_rows)
        miss = sh.get("missing_capabilities") or []
        if miss:
            _sheet(wb, "Битые датчики",
                   ["ТС", "Пропавшие блоки данных"],
                   [[m.get("name") or m.get("terminal_id"),
                     ", ".join(m.get("missing") or [])] for m in miss])

    # титул периода на первом листе
    first = wb.worksheets[0]
    first.insert_rows(1)
    first["A1"] = f"Период: {period}"
    first["A1"].font = Font(bold=True, italic=True, color="6B6B6B")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
