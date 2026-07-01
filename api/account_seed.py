"""Авто-сидинг учёток ДЗО на новые каноничные узлы дерева (при синке).

Оргструктура ресинкается вживую на каждом синке (`build_from_omnicomm_tree`).
Здесь — второй слой той же синхронизации: как только в дереве Omnicomm появляется
НОВЫЙ узел (ДЗО/под-ДЗО) без учётки, ему автоматически заводится логин:
транслит-слаг имени + сгенерированный пароль, роль `manager`, привязка `org_id`.

Инвариант: **идемпотентно и аддитивно** — узел с уже существующей учёткой
пропускается; существующие учётки, пароли и привязки не трогаются; учётки удалённых
узлов НЕ трогаются (остаются в users.json, скоуп пустой — безопасно). Узлы-срезы
(`/…/`) сюда не доходят: `build_from_omnicomm_tree` их уже выкинул.

Пароль в открытом виде существует ТОЛЬКО в `data/accounts.xlsx` (лист раздачи,
gitignored). В `users.json` — лишь PBKDF2-хеш. Новые учётки дописываются в тот же
xlsx строкой (Уровень/Узел/ТС/Логин/Пароль/org_id), чтобы админ/КАП забрал их через
`GET /api/accounts`.
"""

from __future__ import annotations

import os
import re
import secrets
from typing import Iterable, Optional

from omnicomm_report import auth

DEFAULT_XLSX = os.path.join("data", "accounts.xlsx")

# Кириллица → латиница (схема существующих логинов: й→y, ю→yu, я→ya, х→h, ё→e …).
_TRANSLIT = {
    "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ё": "e",
    "ж": "zh", "з": "z", "и": "i", "й": "y", "к": "k", "л": "l", "м": "m",
    "н": "n", "о": "o", "п": "p", "р": "r", "с": "s", "т": "t", "у": "u",
    "ф": "f", "х": "h", "ц": "ts", "ч": "ch", "ш": "sh", "щ": "sch",
    "ъ": "", "ы": "y", "ь": "", "э": "e", "ю": "yu", "я": "ya",
    # казахские
    "ә": "a", "ғ": "g", "қ": "k", "ң": "n", "ө": "o", "ұ": "u", "ү": "u",
    "һ": "h", "і": "i",
}

# Ведущие юр-формы, которые срезаем из слага (как в существующих логинах:
# «ТОО Каратау» → karatau, «АО НАК Казатомпром» → kazatomprom).
_LEGAL_PREFIX = {"тоо", "ао", "оао", "зао", "сп", "нак", "дп", "ак", "рго", "гу"}

# Русские подписи уровня для листа раздачи.
_LEVEL_LABEL = {
    "holding": "Холдинг",
    "dzo": "ДЗО",
    "sub_dzo": "под-ДЗО",
    "contractor": "Подрядчик",
    "unknown": "Узел",
}


def _translit(s: str) -> str:
    out = []
    for ch in (s or "").lower():
        out.append(_TRANSLIT.get(ch, ch))
    return "".join(out)


def slugify_org(name: str) -> str:
    """Логин-слаг из имени организации: срезать ведущие юр-формы, транслитерировать,
    оставить [a-z0-9-]. Пустой результат → 'org'."""
    tokens = re.split(r"\s+", (name or "").strip())
    while tokens and re.sub(r"[^а-яёa-z]", "", tokens[0].lower()) in _LEGAL_PREFIX:
        tokens.pop(0)
    base = _translit(" ".join(tokens) if tokens else (name or "org"))
    base = base.lower().replace(" ", "-")
    base = re.sub(r"[^a-z0-9-]", "", base)
    base = re.sub(r"-{2,}", "-", base).strip("-")
    return base or "org"


def _flatten_kpi(kpi_tree) -> list:
    """Плоский список OrgKPI-узлов дерева (корень + потомки)."""
    out = []

    def walk(nodes):
        for n in nodes:
            out.append(n)
            walk(n.children)

    walk(kpi_tree or [])
    return out


def seed_new_accounts(kpi_tree, *, users_path: str = auth.DEFAULT_USERS_PATH,
                      xlsx_path: str = DEFAULT_XLSX,
                      role: str = "manager") -> list[dict]:
    """Завести учётку на каждый узел дерева без привязанной учётки. Возвращает
    список созданных `{login, password, org_id, node, level, vehicles}`.
    Ничего не удаляет и не переписывает существующее."""
    users = auth._load(users_path)
    have_org = {str(rec.get("org_id")) for rec in users.values() if rec.get("org_id")}
    have_name = set(users.keys())

    created: list[dict] = []
    for node in _flatten_kpi(kpi_tree):
        org = node.org
        oid = str(org.org_id)
        if not oid or oid in have_org:
            continue                      # уже есть учётка на этот узел — пропуск
        base = slugify_org(org.name)
        uname = base if base not in have_name else f"{base}-{oid}"
        while uname in have_name:         # крайне маловероятная вторичная коллизия
            uname = f"{base}-{secrets.token_hex(2)}"
        pw = secrets.token_urlsafe(9)
        if not auth.create_user(uname, pw, role, org_id=oid, path=users_path):
            continue
        have_org.add(oid)
        have_name.add(uname)
        created.append({
            "login": uname, "password": pw, "org_id": oid,
            "node": org.name,
            "level": _LEVEL_LABEL.get(getattr(org.level, "value", str(org.level)), "Узел"),
            "vehicles": int(getattr(node, "kpi", None).vehicles_total
                            if getattr(node, "kpi", None) else 0),
        })

    if created:
        _append_xlsx(xlsx_path, created)
    return created


def _append_xlsx(xlsx_path: str, rows: Iterable[dict]) -> None:
    """Дописать созданные учётки в лист раздачи (создать файл при отсутствии).
    Аддитивно: существующие строки/пароли не трогаем."""
    try:
        from openpyxl import Workbook, load_workbook
    except Exception:  # noqa: BLE001 — без openpyxl xlsx просто не обновится
        return
    os.makedirs(os.path.dirname(xlsx_path) or ".", exist_ok=True)
    header = ("Уровень", "Узел", "ТС", "Логин", "Пароль", "org_id")
    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Учётки ДЗО"
        ws.append(("Учётки доступа к дашборду omnicomm.technokod.kz — "
                   "КОНФИДЕНЦИАЛЬНО, сменить пароли",))
        ws.append(header)
    for r in rows:
        ws.append((r["level"], r["node"], r["vehicles"],
                   r["login"], r["password"], r["org_id"]))
    wb.save(xlsx_path)
    try:
        os.chmod(xlsx_path, 0o600)
    except OSError:
        pass
