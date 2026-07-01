"""Тесты авто-сидинга учёток ДЗО на новые узлы дерева."""

from __future__ import annotations

import os
import sys
from dataclasses import dataclass, field
from typing import Optional

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from api import account_seed  # noqa: E402
from omnicomm_report import auth  # noqa: E402


# --- лёгкие фейки OrgKPI-дерева (нужны только .org/.children/.kpi.vehicles_total) --
@dataclass
class _Lvl:
    value: str


@dataclass
class _Org:
    org_id: str
    name: str
    level: _Lvl


@dataclass
class _Kpi:
    vehicles_total: int


@dataclass
class _Node:
    org: _Org
    kpi: _Kpi
    children: list = field(default_factory=list)


def _node(oid, name, veh, level="dzo", children=None):
    return _Node(_Org(oid, name, _Lvl(level)), _Kpi(veh), children or [])


def _tree():
    return [
        _node("holding", "КАП", 100, "holding", children=[
            _node("8894", "АО НАК Казатомпром", 100, "dzo", children=[
                _node("13766", "ТОО Каратау", 44),
                _node("8922", "ТОО СП ЮГХК", 53),
                _node("13775", "ТОО Уранэнерго ШФУ", 12, "sub_dzo"),
            ]),
        ]),
    ]


# --- слаги ---------------------------------------------------------------------
def test_slugify_strips_legal_forms_and_translit():
    assert account_seed.slugify_org("ТОО Каратау") == "karatau"
    assert account_seed.slugify_org("ТОО СП ЮГХК") == "yughk"
    assert account_seed.slugify_org("ТОО Уранэнерго ШФУ") == "uranenergo-shfu"
    assert account_seed.slugify_org("АО НАК Казатомпром") == "kazatomprom"
    assert account_seed.slugify_org("ТОО Qorgan-Security") == "qorgan-security"
    assert account_seed.slugify_org("АО СП Заречное") == "zarechnoe"
    assert account_seed.slugify_org("ИВТ - ЗЕРДЕ") == "ivt-zerde"


# --- сидинг --------------------------------------------------------------------
def test_seeds_account_per_node(tmp_path):
    up = str(tmp_path / "users.json")
    xp = str(tmp_path / "accounts.xlsx")
    created = account_seed.seed_new_accounts(_tree(), users_path=up, xlsx_path=xp)
    logins = {c["login"] for c in created}
    # 4 узла (холдинг + Казатомпром + Каратау + ЮГХК + ШФУ) = 5
    assert len(created) == 5
    assert {"karatau", "yughk", "uranenergo-shfu"} <= logins
    # каждая учётка привязана к своему org_id и логинится своим паролем
    by_login = {c["login"]: c for c in created}
    assert auth.user_org("karatau", up) == "13766"
    assert auth.authenticate("karatau", by_login["karatau"]["password"], up)["org_id"] == "13766"


def test_seeding_is_idempotent(tmp_path):
    up = str(tmp_path / "users.json")
    xp = str(tmp_path / "accounts.xlsx")
    first = account_seed.seed_new_accounts(_tree(), users_path=up, xlsx_path=xp)
    again = account_seed.seed_new_accounts(_tree(), users_path=up, xlsx_path=xp)
    assert first and again == []                       # второй проход — ничего нового
    assert len(auth.list_users(up)) == len(first)      # без дублей


def test_only_new_nodes_seeded_and_existing_untouched(tmp_path):
    up = str(tmp_path / "users.json")
    xp = str(tmp_path / "accounts.xlsx")
    account_seed.seed_new_accounts(_tree(), users_path=up, xlsx_path=xp)
    karatau_hash = auth._load(up)["karatau"]["hash"]
    # добавилось новое ДЗО в дерево
    tree2 = _tree()
    tree2[0].children[0].children.append(_node("99999", "ТОО Новое ДЗО", 7))
    created = account_seed.seed_new_accounts(tree2, users_path=up, xlsx_path=xp)
    assert [c["org_id"] for c in created] == ["99999"]           # только новый узел
    assert auth._load(up)["karatau"]["hash"] == karatau_hash     # старое не тронуто


def test_username_collision_gets_suffix(tmp_path):
    up = str(tmp_path / "users.json")
    xp = str(tmp_path / "accounts.xlsx")
    tree = [_node("holding", "КАП", 1, "holding", children=[
        _node("100", "ТОО Каратау", 1),
        _node("200", "ТОО Каратау", 1),   # то же имя, другой узел
    ])]
    created = account_seed.seed_new_accounts(tree, users_path=up, xlsx_path=xp)
    logins = [c["login"] for c in created if c["org_id"] in ("100", "200")]
    assert "karatau" in logins
    assert "karatau-200" in logins        # второй получил суффикс org_id
    assert len(set(logins)) == 2          # логины уникальны


def test_xlsx_appended(tmp_path):
    up = str(tmp_path / "users.json")
    xp = str(tmp_path / "accounts.xlsx")
    account_seed.seed_new_accounts(_tree(), users_path=up, xlsx_path=xp)
    from openpyxl import load_workbook
    ws = load_workbook(xp).active
    rows = list(ws.iter_rows(values_only=True))
    header = rows[1]
    assert header == ("Уровень", "Узел", "ТС", "Логин", "Пароль", "org_id")
    logins = {r[3] for r in rows[2:]}
    assert {"karatau", "yughk"} <= logins
