"""Тесты арендного модуля: акт наработки × ставка + лист отклонений."""

from __future__ import annotations

import os
import sys
from datetime import datetime, timezone

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))

from omnicomm_report import rental  # noqa: E402
from omnicomm_report.models import (  # noqa: E402
    FleetKPI, FleetReport, ReportPeriod, VehicleMetrics,
)


def _report(vehicles, fund=0.0):
    kpi = FleetKPI(time_fund_hours_per_day=fund)
    return FleetReport(
        period=ReportPeriod(start=datetime(2026, 6, 1, tzinfo=timezone.utc),
                            end=datetime(2026, 6, 11, tzinfo=timezone.utc)),
        client_name="Аренда-Тест", vehicles=vehicles, kpi=kpi)


def test_act_only_vehicles_with_rate():
    vs = [
        VehicleMetrics(vehicle_id="1", name="Экскаватор-1",
                       engine_hours=80.0, engine_idle_hours=20.0),
        VehicleMetrics(vehicle_id="2", name="Без ставки",
                       engine_hours=50.0),
    ]
    act = rental.build_act(_report(vs), {"Экскаватор-1": {"rate_kzt_per_mh": 15000}})
    assert act is not None
    assert len(act["rows"]) == 1
    assert act["rows"][0]["amount_kzt"] == 80 * 15000
    assert act["total_kzt"] == 1_200_000


def test_act_none_without_rates():
    vs = [VehicleMetrics(vehicle_id="1", name="ТС", engine_hours=10.0)]
    assert rental.build_act(_report(vs), {}) is None


def test_act_deviations_overfund_and_idle():
    """Сверх фонда (10 дней × 8 ч = 80 ч) и высокий холостой ход → отклонения."""
    vs = [VehicleMetrics(vehicle_id="1", name="Кран-1",
                         engine_hours=120.0, engine_idle_hours=70.0)]
    act = rental.build_act(_report(vs, fund=8.0),
                           {"Кран-1": {"rate_kzt_per_mh": 10000}})
    text = " ".join(act["deviations"])
    assert "сверх договорного режима" in text
    assert "холостой ход" in text.lower()


def test_act_no_data_vehicle_flagged():
    v = VehicleMetrics(vehicle_id="1", name="Тёмное ТС", engine_hours=0.0)
    v.has_data = False
    act = rental.build_act(_report([v]), {"Тёмное ТС": {"rate_kzt_per_mh": 9000}})
    assert act is None or not act["rows"]   # нет подтверждённой наработки


def test_export_xlsx(tmp_path):
    vs = [VehicleMetrics(vehicle_id="1", name="Погрузчик-1",
                         engine_hours=60.0, engine_idle_hours=10.0)]
    act = rental.build_act(_report(vs), {"Погрузчик-1": {"rate_kzt_per_mh": 8000}})
    path = rental.export_act_xlsx(act, str(tmp_path / "act.xlsx"))
    assert os.path.exists(path)
    from openpyxl import load_workbook
    ws = load_workbook(path).active
    assert "Акт наработки" in ws["A1"].value
