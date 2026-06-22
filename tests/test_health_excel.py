"""Тесты интеграции ТЗ R6/R7/R3.3/R4.4: health-секции, Excel, ADR-смещение."""

import io
import time

import pytest

from omnicomm_report import demo_data, geozones
from omnicomm_report.models import ReportPeriod
from datetime import datetime, timezone

from api import excel, health


@pytest.fixture
def demo_vehicles():
    period = ReportPeriod(
        start=datetime(2026, 6, 20, tzinfo=timezone.utc),
        end=datetime(2026, 6, 22, tzinfo=timezone.utc))
    return demo_data.demo_fleet(period)


# --- R7 Sensor Health --------------------------------------------------------

def test_sensor_health_demo_counts_cover_fleet(demo_vehicles):
    now = int(time.time())
    sh = health.build_sensor_health_demo(demo_vehicles, now=now)
    assert len(sh["terminals"]) == len(demo_vehicles)
    assert sum(sh["counts"].values()) == len(demo_vehicles)
    assert set(sh["counts"]) >= {"online", "stale", "offline"}
    assert sh["level"] == "terminal"


def test_sensor_health_live_marks_missing_fuel():
    # ТС с GPS/оборотами, но без блока топлива → попадает в missing_capabilities.
    rows = [{"consolidatedReport": {
        "vehicleId": 101, "date": 0,
        "mv": {"mileage": 10, "maxSpeed": 60, "worked": 3600},
        "fuel": {}}}]
    activity = [{"id": 101, "dateID": int(time.time() * 1000)}]
    tree = [{"terminal_id": 101, "name": "КАМАЗ-101", "receive_data": True}]
    sh = health.build_sensor_health(activity, rows, tree, now=int(time.time()))
    assert sh["counts"]["online"] == 1
    assert any("Топливо" in " ".join(m["missing"]) for m in sh["missing_capabilities"])


# --- R6 Контроль ТО ----------------------------------------------------------

def test_maintenance_demo_has_statuses(demo_vehicles):
    mt = health.build_maintenance_demo(demo_vehicles)
    assert len(mt["items"]) == len(demo_vehicles)
    assert set(s["status"] for s in mt["items"]) <= {"ok", "ожидается", "просрочено"}
    assert sum(mt["counts"].values()) == len(demo_vehicles)


def test_maintenance_live_overdue_when_worked_exceeds_interval():
    # Стационарная техника (пробег~0, моточасы большие) → план по моточасам,
    # наработка выше интервала → «просрочено».
    rows = [{"consolidatedReport": {
        "vehicleId": 7, "date": 0, "mv": {"worked": 400 * 3600, "mileage": 0}}}]

    class V:  # минимальный stand-in VehicleMetrics
        vehicle_id, name = "7", "Буровая БУ-7"
        engine_hours, mileage_km, max_speed_kmh = 400.0, 0.0, 0.0
        has_data = True

    mt = health.build_maintenance(records=rows, vehicles=[V()])
    statuses = {s["terminal_id"]: s for s in mt["items"]}
    assert statuses["7"]["status"] == "просрочено"


# --- R3.3 Excel --------------------------------------------------------------

def test_excel_workbook_has_expected_sheets(demo_vehicles):
    now = int(time.time())
    snapshot = {
        "period": {"label": "20–22 июня"},
        "orgs": [], "economics": None, "recommendations": [],
        "sensor_health": health.build_sensor_health_demo(demo_vehicles, now=now),
        "maintenance": health.build_maintenance_demo(demo_vehicles),
    }
    data = excel.build_workbook(snapshot)
    assert isinstance(data, bytes) and len(data) > 2000
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data))
    assert "Контроль ТО" in wb.sheetnames
    assert "Качество данных" in wb.sheetnames


# --- R4.4 ADR-смещение -------------------------------------------------------

def test_adr_offset():
    assert geozones.apply_adr_offset(60, False) == 60
    assert geozones.apply_adr_offset(60, True) < 60
    assert geozones.apply_adr_offset(10, True) >= 0  # не уходит в минус
