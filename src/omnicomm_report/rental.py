"""Арендный модуль v1: акт наработки по моточасам (STRATEGY §4.3).

Боль рынка аренды спецтехники — биллинг по моточасам: ручные счётчики
скручиваются, наработка приписывается. Телематика выступает «нотариусом»:
акт = независимая наработка ТС за период × ставка аренды (₸/мч).

Ставка хранится в паспорте ТС (norms.py, поле `rate_kzt_per_mh`) — enter-once.
ТС без ставки в акт не попадают (это не ошибка: не вся техника в аренде).
Лист отклонений: работа сверх нормативного фонда клиента (вторая смена) и
высокая доля холостого хода — повод для разговора с арендатором.

Выход: dict для рендера + export_act_xlsx() — самостоятельный документ.
"""

from __future__ import annotations

import logging
import os
from typing import Optional

from omnicomm_report import config
from omnicomm_report.models import FleetReport

log = logging.getLogger(__name__)


def build_act(report: FleetReport, norms_data: dict[str, dict]) -> Optional[dict]:
    """Акт наработки: ТС со ставкой аренды → наработка × ставка.

    Возвращает None, если ни у одного ТС не задана ставка `rate_kzt_per_mh`.
    """
    days = max(1.0, (report.period.end_ts - report.period.start_ts) / 86400.0)
    fund_h = report.kpi.time_fund_hours_per_day or 0.0
    rows: list[dict] = []
    deviations: list[str] = []
    for v in report.vehicles:
        rate = (norms_data.get(v.name) or {}).get("rate_kzt_per_mh")
        try:
            rate = float(rate)
        except (TypeError, ValueError):
            continue
        if rate <= 0:
            continue
        if not v.has_data:
            deviations.append(f"«{v.name}»: нет данных за период — наработка "
                              "не подтверждена телематикой.")
            continue
        hours = float(v.engine_hours or 0.0)
        idle_h = float(v.engine_idle_hours or 0.0)
        amount = round(hours * rate)
        rows.append({
            "name": v.name,
            "engine_hours": round(hours, 1),
            "idle_hours": round(idle_h, 1),
            "rate_kzt_per_mh": rate,
            "amount_kzt": amount,
        })
        if fund_h > 0 and hours > fund_h * days:
            deviations.append(
                f"«{v.name}»: наработка {hours:.1f} мч превышает фонд "
                f"{fund_h * days:.0f} ч за период — работа сверх договорного "
                "режима (вторая смена?).")
        if hours > 0 and idle_h / hours >= config.ALERT_IDLE_SHARE:
            deviations.append(
                f"«{v.name}»: холостой ход {idle_h / hours * 100:.0f}% "
                "наработки — моточасы накручены простоем, требует проверки.")
    if not rows:
        return None
    rows.sort(key=lambda r: r["amount_kzt"], reverse=True)
    return {
        "client": report.client_name,
        "period": report.period.human(),
        "rows": rows,
        "total_hours": round(sum(r["engine_hours"] for r in rows), 1),
        "total_kzt": round(sum(r["amount_kzt"] for r in rows)),
        "deviations": deviations,
    }


def export_act_xlsx(act: dict, out_path: str) -> str:
    """Выгрузить акт наработки в .xlsx (документ для арендатора)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Акт наработки"

    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="2F5C8F")

    ws.append([f"Акт наработки техники — {act['client']}"])
    ws.append([f"Период: {act['period']} · наработка подтверждена телематикой"])
    ws.append([])
    ws["A1"].font = Font(bold=True, size=14)

    headers = ["ТС", "Наработка, мч", "в т.ч. холостой ход, ч",
               "Ставка, ₸/мч", "Сумма, ₸"]
    ws.append(headers)
    hrow = ws.max_row
    for c in range(1, len(headers) + 1):
        cell = ws.cell(hrow, c)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center")

    for r in act["rows"]:
        ws.append([r["name"], r["engine_hours"], r["idle_hours"],
                   r["rate_kzt_per_mh"], r["amount_kzt"]])
    ws.append(["ИТОГО", act["total_hours"], "", "", act["total_kzt"]])
    for c in (1, 2, 5):
        ws.cell(ws.max_row, c).font = Font(bold=True)

    if act["deviations"]:
        ws.append([])
        ws.append(["Отклонения (требуют проверки):"])
        ws.cell(ws.max_row, 1).font = Font(bold=True)
        for d in act["deviations"]:
            ws.append([d])

    widths = [34, 16, 22, 14, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    out_path = os.path.abspath(out_path)
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)
    log.info("Акт наработки сохранён: %s (%d ТС, %s ₸)",
             out_path, len(act["rows"]), f"{act['total_kzt']:,}".replace(",", " "))
    return out_path
